"""
Evaluate extracted typical-human participant information against ground truth.

Compared fields:
    Excel N          <-> typical_human_total
    Excel male       <-> typical_human_male_number
    Excel age range  <-> typical_human_age_range

Accuracy denominators include only rows with valid, non-empty ground-truth
values for the corresponding field.
"""

from __future__ import annotations

import argparse
import csv
import json
import re
from pathlib import Path
from typing import Any, Optional, Tuple

import matplotlib

matplotlib.use("Agg")
import matplotlib.pyplot as plt
from openpyxl import load_workbook


STEP3_DIR = Path("D:/language_template/reviewer/current_data/step3_library")
DEFAULT_GROUNDTRUTH_XLSX = (
    STEP3_DIR / "details_studies_OSF_meteinfo_Groudtruth_LLM_infocheck.xlsx"
)
DEFAULT_EXTRACTION_CSV = (
    STEP3_DIR
    / "method_info_extraction_deepseek_v4_flash_groundtruth"
    / "method_info_table.csv"
)
DEFAULT_OUTPUT_DIR = (
    STEP3_DIR
    / "method_info_extraction_deepseek_v4_flash_groundtruth"
    / "groundtruth_evaluation"
)

FIELD_CONFIG = {
    "typical_human_total": {
        "label": "Total participants",
        "groundtruth_column": "N",
    },
    "typical_human_male_number": {
        "label": "Male participants",
        "groundtruth_column": "male",
    },
    "typical_human_age_range": {
        "label": "Age range",
        "groundtruth_column": "age range",
    },
}


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Evaluate extracted participant fields against Excel ground truth."
    )
    parser.add_argument("--groundtruth-xlsx", default=str(DEFAULT_GROUNDTRUTH_XLSX))
    parser.add_argument("--extraction-csv", default=str(DEFAULT_EXTRACTION_CSV))
    parser.add_argument("--output-dir", default=str(DEFAULT_OUTPUT_DIR))
    return parser.parse_args()


def normalize_pmid(value: Any) -> str:
    if value is None or isinstance(value, bool):
        return ""
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    text = str(value).strip()
    if re.fullmatch(r"\d+\.0+", text):
        return text.split(".", 1)[0]
    match = re.search(r"(?<!\d)(\d{5,10})(?!\d)", text)
    return match.group(1) if match else ""


def normalize_count(value: Any) -> Optional[Any]:
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, int):
        return value
    if isinstance(value, float) and value.is_integer():
        return int(value)
    text = str(value).strip()
    if not text:
        return None
    if re.fullmatch(r"\d+(?:\.0+)?", text):
        return int(float(text))
    if "," in text:
        values = []
        valid_tokens = True
        for token in text.split(","):
            token = token.strip()
            if not token or token.lower() in {
                "unknown",
                "unclear",
                "not reported",
                "not specified",
                "n/a",
                "na",
            }:
                continue
            if re.fullmatch(r"\d+(?:\.0+)?", token):
                values.append(int(float(token)))
            else:
                valid_tokens = False
                break
        if valid_tokens and values:
            return tuple(values)
    return None


def format_number(value: float) -> str:
    if float(value).is_integer():
        return str(int(value))
    return f"{value:.4f}".rstrip("0").rstrip(".")


def normalize_age_ranges(value: Any) -> Optional[Tuple[Tuple[str, str], ...]]:
    if value is None or isinstance(value, bool):
        return None

    if isinstance(value, float):
        if value.is_integer():
            return ((str(int(value)), str(int(value))),)
        text = f"{value:.10f}".rstrip("0")
        integer_part, decimal_part = text.split(".", 1)
        # Excel sometimes converts a text range such as 18-35 into 18.35.
        if len(decimal_part) == 2 and integer_part.isdigit():
            return ((integer_part, decimal_part),)
        return ((format_number(value), format_number(value)),)

    if isinstance(value, int):
        text = str(value)
        return ((text, text),)

    text = str(value).strip().lower()
    if not text:
        return None
    text = (
        text.replace("–", "-")
        .replace("—", "-")
        .replace("−", "-")
        .replace("years", "")
        .replace("year", "")
        .replace("yrs", "")
        .replace("yr", "")
    )
    # Decimal commas occur in values such as 19,8-30,8.
    if re.fullmatch(r"\s*\d+,\d+\s*-\s*\d+,\d+\s*", text):
        text = re.sub(r"(?<=\d),(?=\d)", ".", text)
    text = re.sub(r"\bto\b", "-", text)

    pairs = re.findall(
        r"(\d+(?:\.\d+)?)\s*-\s*(\d+(?:\.\d+)?)",
        text,
    )
    if pairs:
        normalized = [
            (format_number(float(low)), format_number(float(high)))
            for low, high in pairs
        ]
        return tuple(sorted(normalized))

    numbers = re.findall(r"\d+(?:\.\d+)?", text)
    if len(numbers) == 1:
        number = format_number(float(numbers[0]))
        return ((number, number),)
    return None


def display_normalized(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, tuple):
        if value and isinstance(value[0], tuple):
            return "; ".join(f"{low}-{high}" for low, high in value)
        return ", ".join(str(item) for item in value)
    return str(value)


def normalized_values_match(
    groundtruth_value: Any,
    extracted_value: Any,
) -> bool:
    """Match one ground-truth study value against ordered experiment values."""

    if groundtruth_value == extracted_value:
        return True
    if isinstance(extracted_value, tuple):
        if isinstance(groundtruth_value, tuple):
            if len(groundtruth_value) == 1:
                return groundtruth_value[0] in extracted_value
        else:
            return groundtruth_value in extracted_value
    return False


def load_groundtruth(path: Path) -> dict[str, dict[str, Any]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    results: dict[str, dict[str, Any]] = {}
    for worksheet in workbook.worksheets:
        header = next(
            worksheet.iter_rows(min_row=1, max_row=1, values_only=True)
        )
        columns = {
            str(value or "").strip().lower(): index
            for index, value in enumerate(header)
            if str(value or "").strip()
        }
        required = ["pmid", "n", "male", "age range"]
        if not all(name in columns for name in required):
            continue
        for row in worksheet.iter_rows(min_row=2, values_only=True):
            pmid = normalize_pmid(row[columns["pmid"]])
            if not pmid:
                continue
            results[pmid] = {
                "N": row[columns["n"]],
                "male": row[columns["male"]],
                "age range": row[columns["age range"]],
            }
    return results


def load_extractions(path: Path) -> dict[str, dict[str, str]]:
    with path.open("r", encoding="utf-8-sig", newline="") as file_obj:
        return {
            str(row.get("PMID", "")).strip(): row
            for row in csv.DictReader(file_obj)
            if str(row.get("PMID", "")).strip()
        }


def main() -> None:
    args = parse_args()
    groundtruth_path = Path(args.groundtruth_xlsx)
    extraction_path = Path(args.extraction_csv)
    output_dir = Path(args.output_dir)
    if not groundtruth_path.exists():
        raise FileNotFoundError(groundtruth_path)
    if not extraction_path.exists():
        raise FileNotFoundError(extraction_path)
    output_dir.mkdir(parents=True, exist_ok=True)

    groundtruth = load_groundtruth(groundtruth_path)
    extractions = load_extractions(extraction_path)
    common_pmids = sorted(
        set(groundtruth) & set(extractions),
        key=lambda value: (not value.isdigit(), int(value) if value.isdigit() else value),
    )

    detail_rows = []
    mismatches_by_pmid: dict[str, Dict[str, Any]] = {}
    stats = {
        field: {
            "label": config["label"],
            "eligible": 0,
            "correct": 0,
            "incorrect": 0,
            "groundtruth_missing_or_invalid": 0,
            "extraction_missing": 0,
        }
        for field, config in FIELD_CONFIG.items()
    }

    for pmid in common_pmids:
        detail = {"PMID": pmid}
        for field, config in FIELD_CONFIG.items():
            gt_raw = groundtruth[pmid].get(config["groundtruth_column"])
            extracted_raw = extractions[pmid].get(field, "")
            if field == "typical_human_age_range":
                gt_normalized = normalize_age_ranges(gt_raw)
                extracted_normalized = normalize_age_ranges(extracted_raw)
            else:
                gt_normalized = normalize_count(gt_raw)
                extracted_normalized = normalize_count(extracted_raw)

            if gt_normalized is None:
                status = "groundtruth_missing_or_invalid"
                stats[field]["groundtruth_missing_or_invalid"] += 1
            else:
                stats[field]["eligible"] += 1
                if extracted_normalized is None:
                    status = "incorrect"
                    stats[field]["incorrect"] += 1
                    stats[field]["extraction_missing"] += 1
                elif normalized_values_match(
                    gt_normalized,
                    extracted_normalized,
                ):
                    status = "correct"
                    stats[field]["correct"] += 1
                else:
                    status = "incorrect"
                    stats[field]["incorrect"] += 1

            if status == "incorrect":
                mismatch = mismatches_by_pmid.setdefault(
                    pmid,
                    {
                        "PMID": pmid,
                        "mismatched_fields": {},
                        "structured_content_file": str(
                            STEP3_DIR
                            / "final_segmented_for_analysis"
                            / "full_segmented"
                            / f"paper_{pmid}_structured_content.json"
                        ),
                        "method_info_file": str(
                            STEP3_DIR
                            / "method_info_extraction_deepseek_v4_flash_groundtruth"
                            / "json"
                            / f"paper_{pmid}_method_info.json"
                        ),
                    },
                )
                mismatch["mismatched_fields"][field] = {
                    "label": config["label"],
                    "groundtruth_raw": gt_raw,
                    "extracted_raw": extracted_raw,
                    "groundtruth_normalized": display_normalized(gt_normalized),
                    "extracted_normalized": display_normalized(
                        extracted_normalized
                    ),
                }

            prefix = field.replace("typical_human_", "")
            detail[f"{prefix}_groundtruth_raw"] = gt_raw
            detail[f"{prefix}_extracted_raw"] = extracted_raw
            detail[f"{prefix}_groundtruth_normalized"] = display_normalized(
                gt_normalized
            )
            detail[f"{prefix}_extracted_normalized"] = display_normalized(
                extracted_normalized
            )
            detail[f"{prefix}_status"] = status
        detail_rows.append(detail)

    summary_rows = []
    for field, values in stats.items():
        eligible = values["eligible"]
        accuracy = values["correct"] / eligible if eligible else 0.0
        values["accuracy"] = accuracy
        summary_rows.append(
            {
                "field": field,
                "label": values["label"],
                "correct": values["correct"],
                "incorrect": values["incorrect"],
                "eligible_groundtruth": eligible,
                "accuracy": round(accuracy, 6),
                "accuracy_percent": round(accuracy * 100, 2),
                "groundtruth_missing_or_invalid": values[
                    "groundtruth_missing_or_invalid"
                ],
                "extraction_missing": values["extraction_missing"],
            }
        )

    detail_path = output_dir / "participant_field_comparison.csv"
    with detail_path.open("w", encoding="utf-8-sig", newline="") as file_obj:
        writer = csv.DictWriter(file_obj, fieldnames=list(detail_rows[0]))
        writer.writeheader()
        writer.writerows(detail_rows)

    summary_path = output_dir / "participant_field_accuracy.csv"
    with summary_path.open("w", encoding="utf-8-sig", newline="") as file_obj:
        writer = csv.DictWriter(file_obj, fieldnames=list(summary_rows[0]))
        writer.writeheader()
        writer.writerows(summary_rows)

    report = {
        "groundtruth_unique_pmids": len(groundtruth),
        "extracted_pmids": len(extractions),
        "common_pmids": len(common_pmids),
        "accuracy_definition": (
            "Exact match after field-specific normalization; denominator includes "
            "only valid, non-empty ground-truth values."
        ),
        "fields": stats,
    }
    report_path = output_dir / "participant_field_accuracy.json"
    report_path.write_text(
        json.dumps(report, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )

    mismatch_records = sorted(
        mismatches_by_pmid.values(),
        key=lambda row: (
            not str(row["PMID"]).isdigit(),
            int(row["PMID"]) if str(row["PMID"]).isdigit() else row["PMID"],
        ),
    )
    mismatch_report = {
        "mismatched_paper_count": len(mismatch_records),
        "evaluated_common_pmids": len(common_pmids),
        "definition": (
            "Only fields with valid ground-truth values that disagree after "
            "normalization are included."
        ),
        "papers": mismatch_records,
    }
    mismatch_path = output_dir / "mismatched_pmids.json"
    mismatch_path.write_text(
        json.dumps(mismatch_report, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )

    labels = [row["label"] for row in summary_rows]
    percentages = [row["accuracy_percent"] for row in summary_rows]
    colors = ["#2878B5", "#3BA272", "#E69F00"]
    fig, ax = plt.subplots(figsize=(9, 5.6), dpi=180)
    bars = ax.bar(labels, percentages, color=colors, width=0.58)
    ax.set_ylim(0, 100)
    ax.set_ylabel("Accuracy (%)")
    ax.set_title("DeepSeek Method Information vs. Ground Truth")
    ax.grid(axis="y", linestyle="--", linewidth=0.7, alpha=0.35)
    ax.spines["top"].set_visible(False)
    ax.spines["right"].set_visible(False)
    for bar, row in zip(bars, summary_rows):
        ax.text(
            bar.get_x() + bar.get_width() / 2,
            bar.get_height() + 2,
            f"{row['accuracy_percent']:.1f}%\n"
            f"({row['correct']}/{row['eligible_groundtruth']})",
            ha="center",
            va="bottom",
            fontsize=10,
        )
    fig.tight_layout()
    chart_path = output_dir / "participant_field_accuracy.png"
    fig.savefig(chart_path, bbox_inches="tight")
    plt.close(fig)

    print(f"Common PMIDs: {len(common_pmids)}")
    for row in summary_rows:
        print(
            f"{row['label']}: {row['accuracy_percent']:.2f}% "
            f"({row['correct']}/{row['eligible_groundtruth']})"
        )
    print(f"Details: {detail_path}")
    print(f"Summary: {summary_path}")
    print(f"Chart: {chart_path}")
    print(f"Mismatched PMIDs: {mismatch_path}")


if __name__ == "__main__":
    main()
